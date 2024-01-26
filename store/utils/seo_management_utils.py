import io
import os

import openpyxl
from django.conf import settings
from django.core.management import call_command
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import render
from django.utils.http import quote
from django.views import View
from openpyxl.styles import Font

from seo_manager.models import SEOAttributes


def export_seo_to_xlsx(request):
    print('Export start')

    # Making .xlsx file
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Seo Attrs'

    headers = ['page_url', 'title', 'meta_description', 'alt_image']
    for col, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    row = 2

    for page_data in SEOAttributes.objects.all():
        seo_data = (
            page_data.page_url,
            page_data.title,
            page_data.meta_description,
            page_data.alt_image,
        )

        for col, data in enumerate(seo_data, start=1):
            cell = worksheet.cell(row=row, column=col, value=data)

        row += 1

    output_file = io.BytesIO()
    workbook.save(output_file)
    output_file.seek(0)
    response = HttpResponse(output_file.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{quote("seo_attrs.xlsx")}"'

    return response


def import_seo_from_xlsx(request):
    if request.method == 'POST':
        file = request.FILES.get('file')
        if file:
            print('Import start')
            file_path = os.path.join(settings.BASE_DIR, 'feeds/seo/import', file.name)

            with open(file_path, 'wb') as destination:
                for chunk in file.chunks():
                    destination.write(chunk)

            workbook = openpyxl.load_workbook(file_path)
            worksheet = workbook.active

            with transaction.atomic():
                for row in worksheet.iter_rows(min_row=2, values_only=True):
                    seo_data = {
                        'page_url': row[0],
                        'title': row[1] if row[1] else '',
                        'meta_description': row[2] if row[2] else '',
                        'alt_image': row[3] if row[3] else '',
                    }
                    SEOAttributes.objects.update_or_create(page_url=row[0], defaults=seo_data)

            workbook.close()
            os.remove(file_path)

            return HttpResponse("Import SEO complete")

    return render(request, 'admin/import_seo_form.html')


class CreateSEOAttributesFromSitemapView(View):
    def get(self, request, *args, **kwargs):
        try:
            call_command('create_seo_attributes')
            return HttpResponse("SEOAttributes created successfully.")
        except Exception as e:
            return HttpResponse(f"An error occurred: {str(e)}", status=500)
