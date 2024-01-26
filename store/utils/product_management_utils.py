import io
import os

import openpyxl
from django.conf import settings
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import render
from django.utils.http import quote
from openpyxl.styles import Font

from products.models import (Manufacturer, Product, ProductCategory,
                             ProductSubCategory, ProductCharacteristic)


def export_product_to_xlsx(request):
    print('Export start')

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Product Cards'

    headers = ['ID', 'Name', 'Manufacturer', 'Description', 'Price', 'Category', 'Subcategory', 'Discount Percentage',
               'discount_price', 'total_price', 'Article Number', 'Quantity', 'Image', 'Stripe Product Price ID']
    for col, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    row = 2  # Start

    for product in Product.objects.all().select_related('category', 'sub_category'):
        product_data = (
            product.id,
            product.name,
            product.manufacturer.name,
            product.description,
            product.price,
            product.category.name,
            product.sub_category.name,
            product.discount_percentage,
            product.discount_price,
            product.total_price,
            product.article_number,
            product.quantity,
            product.image.url if product.image else "",
            product.stripe_product_price_id,
        )

        for col, data in enumerate(product_data, start=1):
            cell = worksheet.cell(row=row, column=col, value=data)

        row += 1

    output_file = io.BytesIO()
    workbook.save(output_file)
    output_file.seek(0)

    response = HttpResponse(output_file.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{quote("product_cards.xlsx")}"'

    return response


def import_products_from_xlsx(request):
    if request.method == 'POST':
        file = request.FILES.get('file')
        if file:
            print('Import start')
            file_path = os.path.join(settings.BASE_DIR, 'feeds', 'products', 'import', file.name)

            try:
                with open(file_path, 'wb') as destination:
                    for chunk in file.chunks():
                        destination.write(chunk)

                workbook = openpyxl.load_workbook(file_path)
                worksheet = workbook.active

                with transaction.atomic():
                    for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                        product_data = {
                            'name': row[0],
                            'manufacturer': row[1],
                            'description': row[2],
                        }

                        price_str = row[3]
                        if price_str:
                            try:
                                price = float(price_str)
                            except ValueError:
                                price = 1000000
                        else:
                            price = 1000000

                        product_data['price'] = price

                        manufacturer = row[1]
                        if manufacturer:
                            try:
                                manufacturer = Manufacturer.objects.get(name=manufacturer)
                            except Manufacturer.DoesNotExist:
                                manufacturer = Manufacturer.objects.create(name=manufacturer)
                        else:
                            manufacturer = None

                        category_name = row[4]
                        if not category_name:
                            break
                        try:
                            category = ProductCategory.objects.get(name=category_name)
                        except ProductCategory.DoesNotExist:
                            category = ProductCategory.objects.create(name=category_name)

                        subcategory_name = row[5]
                        try:
                            subcategory = ProductSubCategory.objects.get(name=subcategory_name, parent_category=category)
                        except ProductSubCategory.DoesNotExist:
                            subcategory = ProductSubCategory.objects.create(name=subcategory_name, parent_category=category)

                        product_data['manufacturer'] = manufacturer
                        product_data['category'] = category
                        product_data['sub_category'] = subcategory

                        image_filename = row[7]

                        if image_filename:
                            image_path = os.path.join('products_images', image_filename)
                            product_data['image'] = image_path

                        product, created = Product.objects.update_or_create(name=row[0], defaults=product_data)

                        tth_text = row[6]
                        if tth_text:
                            tth_parts = tth_text.split('***')
                            tth_dict = {}

                            for part in tth_parts:
                                if '---' in part:
                                    key, value = part.split('---')
                                    tth_dict[key] = value
                                else:
                                    pass
                            for name, value in tth_dict.items():
                                characteristic, created = ProductCharacteristic.objects.get_or_create(
                                    product=product, name=name, defaults={'value': value}
                                )
                        print(f'Продукт обработан:{row[0]}')

                workbook.close()
                os.remove(file_path)

                return HttpResponse("Import complete")

            except Exception as e:
                error_message = f"ОШИБКА ПРОИЗОШЛА ПРИ ОБРАБОТКЕ СТРОКИ ТАБЛИЦЫ - {row_number}: {str(e)}"
                return HttpResponse(error_message)

    return render(request, 'admin/import_products_form.html')


def transfer_products_between_manufacturers(request):
    manufacturers = Manufacturer.objects.all()

    if request.method == 'POST':
        from_manufacturer_id = request.POST.get('from_manufacturer')
        to_manufacturer_id = request.POST.get('to_manufacturer')

        if from_manufacturer_id and to_manufacturer_id:
            with transaction.atomic():
                from_manufacturer = Manufacturer.objects.get(pk=from_manufacturer_id)
                to_manufacturer = Manufacturer.objects.get(pk=to_manufacturer_id)

                products_to_transfer = Product.objects.filter(manufacturer=from_manufacturer)
                products_to_transfer.update(manufacturer=to_manufacturer)

                return HttpResponse("Трансфер товаров выполнен успешно.")

    return render(request, 'admin/transfer_products_between_manufacturers.html', {'manufacturers': manufacturers})